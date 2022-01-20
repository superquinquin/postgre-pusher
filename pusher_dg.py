import os
import re
import sys
import csv
import datetime
import openpyxl
import traceback
import unicodedata

import psycopg2
import dropbox

from keys import key







class puller:

  def __init__(self, target_path, dbx_path, token):
    try:
      self.target_path = target_path
      self.dbx_root = dbx_path

      self.dbx_token = token

      self.date = datetime.datetime.now()

      self.dbx_files_map = []
      self.dbx_removed = []
      self.target_files_map = []
      self.table_to_drop = []

      self.log = open(os.path.dirname(__file__)+'/log.txt','a', encoding='utf-8', errors='ignore')
      self.log.write('--------------------------\n')
      self.write_log('init puller', False,'', '')

    except Exception as e:
      self.write_log('init error', True, e, '')
  


  def dbx_connection(self):
    try:
      dbx = dropbox.Dropbox(self.dbx_token)

      self.write_log('dropbox connection done', False,'', '')

    except Exception as e:
      self.write_log('dropbox connection error', True, e, '')
    
    return dbx
  


  def write_log(self, msg, trace_back, e, add):
    self.log.write('--'+datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")+'-- '+msg+' -- '+add+'\n')

    if trace_back == True:
      self.log.write('--'+datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")+'-- '+str(e)+'\n'+str(traceback.extract_tb(sys.exc_info()[2]))+'\n')


  def map_target_folder(self):
    """map the csv and xlsx files in the target folder/subfolders
    reurn list of tuple (name,path)"""
    for (dirpath, dirnames, filenames) in os.walk(self.target_path):
      self.target_files_map += [(file, os.path.join(dirpath, file)) for file in filenames if file.split('.')[-1] in ['csv', 'xlsx']]



  def fetch_table(self, dbx, dbx_path):
    """ dump recently modified csv and xlsx documents and map files in dropbox as list of tuple (name,path)"""

    for file in dbx.files_list_folder(dbx_path).entries:
      name = file.name
      path = file.path_lower

      try:
        last_mod = file.server_modified
        delta = int((self.date - last_mod).total_seconds())

      except AttributeError as e:
        last_mod = None


      if len(name.split('.')) == 1: #recursion
        sub_dbx_path = path
        self.fetch_table(dbx, sub_dbx_path)
      
      else:
        extension = name.split('.')[-1]

        if extension in ['csv', 'xlsx']:
          try:
            self.dbx_files_map.append((name, path))

            if delta < 86400:
              dbx.files_download_to_file(self.target_path+'/'+name, path)

              self.write_log(name, False, '', 'dump sucessfully')

          except Exception as e:
            self.write_log(name, True, e, 'dump error')
  


  def remove_from_target_folder(self):
    in_dbx = [x[0] for x in self.dbx_files_map]
    for file in self.target_files_map:
      if file and file[0] not in in_dbx:
        try:
          self.table_to_drop.append(file)
          os.remove(file[1])

          self.write_log('removing tables from the folder', False,'', str(self.table_to_drop))

        except Exception as e:
          self.write_log('removing table error', True, e, '')


  def run(self, path):
    self.map_target_folder()
    dbx = self.dbx_connection()
    self.fetch_table(dbx, path)
    self.remove_from_target_folder()









class pusher:
  def __init__(self, target_path, table_to_drop, enable_drop):
    try:
      # self.path = '/'.join(os.path.dirname(__file__).split('/')[:-1])
      self.path = target_path
      self.table_to_drop = table_to_drop
      self.enable_drop_table = enable_drop
      self.log = open(os.path.dirname(__file__)+'/log.txt','a', encoding='utf-8', errors='ignore')
      self.table_list = []

      self.log.write('--------------------------\n')
      self.write_log('init pusher', False,'', '')

    except Exception as e:
      self.write_log('init error', True, e, '')



  def connect(self, database, user, password, host, port):
    try:
      self.conn = psycopg2.connect(host=host,
                                   port=port,
                                   database=database,
                                   user=user,
                                   password=password)

      self.cursor = self.conn.cursor()
      
      self.write_log('connection done', False,'', '')

    except Exception as e:
      self.write_log('connection error', True, e, '')



  def close(self):
    try:
      self.log.close()
      self.conn.close()

    except Exception:
      pass

  def write_log(self, msg, trace_back, e, add):
    self.log.write('--'+datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")+'-- '+msg+' -- '+add+'\n')

    if trace_back == True:
      self.log.write('--'+datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")+'-- '+str(e)+'\n'+str(traceback.extract_tb(sys.exc_info()[2]))+'\n')




  def table_set(self):
    try:
      table_query = "SELECT * FROM information_schema.tables;"
      self.cursor.execute(table_query)
      self.table_list = list(self.cursor.fetchall())

    except Exception as e:
      self.write_log('fetch error', True, e, '')


  def name_cleaner(self, name):
    name = ''.join([x for x in  unicodedata.normalize('NFD', name) if unicodedata.category(x) != 'Mn']).lower()
    name = re.sub(r'[^.a-zA-Z0-9]', "_", name)
    name = re.sub(r'_{2,}','_', name)
    
    return name


  def read_csv(self, path):
    doc = open(path, encoding='utf-8', errors='ignore')
    csvfile = csv.reader(doc, delimiter=',')

    columns, content = [], []
    for n,row in enumerate(csvfile):
      if n == 0:
        columns.extend(row)
      else:
        content.append(row)

    return columns, content


  def read_excel(self, path):
    doc = openpyxl.load_workbook(path).active

    columns, content = [], []
    rows, cols = doc.max_row, doc.max_column

    for row in range(1,rows+1):
      r = []
      for col in range(1, cols+1):
        r.append(doc.cell(row = row, column = col).value)
      
      if row == 1:
        columns.extend(r)
      else:
        content.append(r)
    return columns, content



  def correct_bool(self, content, columns):
    match_bool = re.compile(r'vrai|faux|true|false')

    for idx in range(len(columns)):
      col = [x[idx] for x in content]
      uniq = list(set(col))
      
      matched = list(filter(None,[re.findall(match_bool, x.lower())  for x in uniq if type(x) == str]))

      if (len(uniq) == 2 and len(matched) == 2) or (len(uniq) == 1 and len(matched) == 1):
        for row in content:

          if row[idx].lower() in ['vrai','true']:
            row[idx] = True

          else:
            row[idx] = False
            
    return content


  def csv_preprocessing(self, path, file):
    name = self.name_cleaner(file.split('.')[0])
    columns, content = self.read_csv(path+'/'+file)
    content = self.correct_bool(content, columns)

    return name, columns, content



  def xlsx_preprocessing(self, path, file):
    name = self.name_cleaner(file.split('.')[0])
    columns, content = self.read_excel(path+'/'+file)
    content = self.correct_bool(content, columns)
    
    return name, columns, content



  def label_data_type(self, columns, content):
    match1 = re.compile(r'\d+-\d+-\d+')
    match2 = re.compile(r'(\d+/\d+/\d+)')
    match3 = re.compile(r'(\d+\.\d+\.\d+)')
    match_bool = re.compile(r'vrai|faux|true|false')

    sql_cols_type, sql_val_format, sql_cols = [], [], []
    for idx in range(len(columns)):

      col_name = self.name_cleaner(columns[idx])
      sql_cols.append(col_name)

      col = [x[idx] for x in content]

      atype = None
      for i in col:
        if i is not  None or (type(i)== str and i.lower() != 'none'):
          atype = type(i)
          break


      if atype == datetime.datetime:
        sql_val_format.append('%s')
        sql_cols_type.append(col_name + ' DATE')

      elif atype == str:
        for i in col:
          if i is not  None or (type(i)== str and i.lower() != 'none'):
            exmpl = i
            break

        if re.findall(match1,exmpl) or re.findall(match2,exmpl) or re.findall(match3,exmpl):
          sql_cols_type.append(col_name + ' DATE')
          sql_val_format.append('%s')

        elif re.findall(match_bool, exmpl.lower()):
          sql_cols_type.append(col_name + ' BOOL')
          sql_val_format.append('%s')
        
        else:
          sql_cols_type.append(col_name + ' TEXT')
          sql_val_format.append('%s')        


      elif atype == int:
        sql_cols_type.append(col_name + ' INT8')
        sql_val_format.append('%s')

      elif atype == float:
        sql_cols_type.append(col_name + ' FLOAT8')
        sql_val_format.append('%s')

      elif atype == bool:
        sql_cols_type.append(col_name + ' BOOL')
        sql_val_format.append('%s')

      else:
        sql_cols_type.append(col_name + ' TEXT')
        sql_val_format.append('%s')


    sql_val_format = ', '.join(sql_val_format)
    sql_cols_type = ', '.join(sql_cols_type)
    sql_cols = ', '.join(sql_cols)

    return sql_cols_type, sql_val_format, sql_cols



  def search_table(self, name):
    in_place = False
    if name in [x[2] for x in self.table_list]:
      in_place = True

    return in_place



  def query_create_table(self,name, sql_cols_type):
    create_query = "CREATE TABLE IF NOT EXISTS "+name+" ("+sql_cols_type+")"
    self.cursor.execute(create_query)
    self.conn.commit()

  def query_delete_content(self, name):
    delete_query = "DELETE FROM "+name
    self.cursor.execute(delete_query)
    self.conn.commit()


  def query_drop_table(self, name):
    drop_query = "DROP TABLE "+name
    self.cursor.execute(drop_query)
    self.conn.commit()


  def query_push_content(self, name, sql_cols, sql_val_format, content):
    push_query = "INSERT INTO "+name+" ("+sql_cols+") VALUES ("+sql_val_format+")"
    self.cursor.executemany(push_query, content)
    self.conn.commit()



  def launch_payload(self, path):
    architecture = os.listdir(path)
    
    for file in architecture:

      if len(file.split('.')) == 1: 
        sub_path = path+'/'+file
        self.launch_payload(sub_path)

      else:
        extension = file.split('.')[1]

        if extension == 'csv':
          try:
            name, df, content = self.csv_preprocessing(path, file)
            sql_cols_type, sql_val_format, sql_cols = self.label_data_type(df)
            in_place = self.search_table(name)

            if in_place == True:
              self.query_drop_table(name)
              self.query_create_table(name, sql_cols_type)
              self.query_push_content(name, sql_cols, sql_val_format, content)

            else:
              self.query_create_table(name, sql_cols_type)
              self.query_push_content(name, sql_cols, sql_val_format, content)

            self.write_log(file, False, '', 'pushed sucessfully')

          except Exception as e:
            self.write_log('payload error', True, e, '')


        elif extension == 'xlsx':
          try:
            name, columns, content = self.xlsx_preprocessing(path, file)
            sql_cols_type, sql_val_format, sql_cols = self.label_data_type(columns, content)
            in_place = self.search_table(name)

            if in_place == True:
              self.query_drop_table(name)
              self.query_create_table(name, sql_cols_type)
              self.query_push_content(name, sql_cols, sql_val_format, content)

            else:
              self.query_create_table(name, sql_cols_type)
              self.query_push_content(name, sql_cols, sql_val_format, content)

            self.write_log(file, False, '', 'pushed sucessfully')

          except Exception as e:
            self.write_log('payload error', True, e, '') 
  
  def drop_removed_tables(self):
    for table in self.table_to_drop:
      try:
        name = table[0].split('.')[0]
        name = self.name_cleaner(name)
        self.query_drop_table(name)

        self.write_log(name, False, '', 'droped sucessfully')

      except Exception as e:
        self.write_log('droping removed table error', True, e, '')
      

  def run(self, database, username, password, host, port):
    self.connect(database, 
                 username,
                 password,
                 host,
                 port)
    self.table_set()
    self.launch_payload(self.path)

    if self.enable_drop_table == True:
      self.drop_removed_tables()

    self.close()





if __name__ == '__main__':

  pull = puller(key['path'],
                key['dbx_path'],
                key['token'])

  pull.run(key['dbx_path'])

  odoo_pusher = pusher(key['path'], pull.table_to_drop, True).run(key['database'], 
                                                                  key['username'],
                                                                  key['password'],
                                                                  key['host'],
                                                                  key['port'])